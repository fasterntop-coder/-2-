#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from mode1_2352 import RAW_SECTOR_SIZE, verify_mode1_sector

DISC_SIZE = 659_293_824
USER_OFF = 16
USER_SIZE = 2048
B118_DISC_SHA = "75f300e59bd3ad63ca11d4981f328107aa59397fa894abbf5d02476a6457df20"
B118_WORKBOOK_SHA = "e8c85862c10b6d30ed21156b17ca93be834c5cb5f76cf1f58d97c1db6ca22ce9"
B284_STATUS = "PASS_BATCH284_EVENT42_LATE_TAIL_PHYSICAL_PROMOTION"


def sha_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while chunk := f.read(8 * 1024 * 1024):
            h.update(chunk)
    return h.hexdigest()


def extract_asset(raw: bytes | bytearray, lba: int, size: int) -> bytes:
    out = bytearray()
    idx = 0
    while len(out) < size:
        off = (lba + idx) * RAW_SECTOR_SIZE
        sec = raw[off:off + RAW_SECTOR_SIZE]
        if len(sec) != RAW_SECTOR_SIZE or sec[15] != 1:
            raise SystemExit(f"FAIL non-MODE1 sector while extracting asset at LBA {lba + idx}")
        take = min(USER_SIZE, size - len(out))
        out += sec[USER_OFF:USER_OFF + take]
        idx += 1
    return bytes(out)


def load_workbook_tables(path: Path) -> tuple[list[dict], list[dict]]:
    if sha_file(path) != B118_WORKBOOK_SHA:
        raise SystemExit("FAIL B118 workbook SHA-256")
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise SystemExit("FAIL openpyxl required: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)

    def rows_as_dicts(sheet_name: str) -> list[dict]:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"FAIL missing workbook sheet {sheet_name!r}")
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(x) if x is not None else "" for x in next(it)]
        except StopIteration:
            raise SystemExit(f"FAIL empty workbook sheet {sheet_name!r}")
        out = []
        for row in it:
            if not row or all(v is None for v in row):
                continue
            out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
        return out

    sectors = rows_as_dicts("Sector Audit 1626")
    assets = rows_as_dicts("Assets 58")
    if len(sectors) != 1626:
        raise SystemExit(f"FAIL sector audit row count {len(sectors)} != 1626")
    if len(assets) != 58:
        raise SystemExit(f"FAIL asset row count {len(assets)} != 58")
    return sectors, assets


def main() -> None:
    ap = argparse.ArgumentParser(description="Batch285: union exact B118 static58 raw sectors onto a successful Batch284 Disc 1 candidate")
    ap.add_argument("--parent-bin", required=True, type=Path, help="Batch284 output BIN")
    ap.add_argument("--b284-report", required=True, type=Path, help="Batch284 JSON report produced by promote_batch284_event42_late_tail.py")
    ap.add_argument("--b118-donor-bin", required=True, type=Path, help="Exact historical Batch118 full BIN")
    ap.add_argument("--b118-workbook", required=True, type=Path, help="Exact B118 workbook containing Sector Audit 1626 and Assets 58")
    ap.add_argument("--output-bin", required=True, type=Path)
    ap.add_argument("--report", required=True, type=Path)
    args = ap.parse_args()

    if args.parent_bin.stat().st_size != DISC_SIZE or args.b118_donor_bin.stat().st_size != DISC_SIZE:
        raise SystemExit("FAIL disc size")

    parent_sha = sha_file(args.parent_bin)
    donor_sha = sha_file(args.b118_donor_bin)
    if donor_sha != B118_DISC_SHA:
        raise SystemExit(f"FAIL B118 donor SHA expected={B118_DISC_SHA} actual={donor_sha}")

    b284 = json.loads(args.b284_report.read_text(encoding="utf-8"))
    if b284.get("status") != B284_STATUS:
        raise SystemExit("FAIL B284 report status")
    if str(b284.get("output_sha256", "")).lower() != parent_sha:
        raise SystemExit("FAIL B284 report/output BIN SHA binding")
    if int(b284.get("guessed_payload_bytes", -1)) != 0:
        raise SystemExit("FAIL B284 guessed payload bytes is not zero")

    sector_rows, asset_rows = load_workbook_tables(args.b118_workbook)
    sector_map: dict[int, dict] = {}
    for r in sector_rows:
        try:
            lba = int(r["raw_lba"])
            orig = str(r["expected_original_sha256"]).lower()
            patched = str(r["patched_sha256"]).lower()
        except Exception as e:
            raise SystemExit("FAIL malformed Sector Audit row") from e
        if lba in sector_map:
            raise SystemExit(f"FAIL duplicate Sector Audit LBA {lba}")
        if len(orig) != 64 or len(patched) != 64:
            raise SystemExit(f"FAIL malformed sector SHA at LBA {lba}")
        sector_map[lba] = {"lba": lba, "asset": str(r.get("asset", "")), "original": orig, "patched": patched}
    if len(sector_map) != 1626:
        raise SystemExit("FAIL unique sector audit LBA count != 1626")

    parent = args.parent_bin.read_bytes()
    donor = args.b118_donor_bin.read_bytes()
    out = bytearray(parent)
    expected_write: list[dict] = []
    already_target = 0

    for lba in sorted(sector_map):
        rec = sector_map[lba]
        off = lba * RAW_SECTOR_SIZE
        before = bytes(parent[off:off + RAW_SECTOR_SIZE])
        target = bytes(donor[off:off + RAW_SECTOR_SIZE])
        before_sha = sha_bytes(before)
        target_sha = sha_bytes(target)

        if target_sha != rec["patched"]:
            raise SystemExit(f"FAIL B118 donor sector SHA LBA {lba}")
        if not verify_mode1_sector(target)["valid"]:
            raise SystemExit(f"FAIL B118 donor sector EDC/ECC LBA {lba}")

        if before_sha == rec["patched"]:
            already_target += 1
            continue
        if before_sha != rec["original"]:
            raise SystemExit(
                f"FAIL third variant LBA {lba} asset={rec['asset']} "
                f"parent={before_sha} expected_original={rec['original']} target={rec['patched']}"
            )

        expected_write.append({
            "lba": lba,
            "asset": rec["asset"],
            "before_sha256": before_sha,
            "after_sha256": target_sha
        })

    for w in expected_write:
        off = w["lba"] * RAW_SECTOR_SIZE
        out[off:off + RAW_SECTOR_SIZE] = donor[off:off + RAW_SECTOR_SIZE]

    expected_lbas = [w["lba"] for w in expected_write]
    actual_lbas = []
    for lba in range(DISC_SIZE // RAW_SECTOR_SIZE):
        off = lba * RAW_SECTOR_SIZE
        if parent[off:off + RAW_SECTOR_SIZE] != out[off:off + RAW_SECTOR_SIZE]:
            actual_lbas.append(lba)
    if actual_lbas != expected_lbas:
        raise SystemExit("FAIL actual changed-LBA set != Expected Write LBA set")

    for w in expected_write:
        off = w["lba"] * RAW_SECTOR_SIZE
        sec = bytes(out[off:off + RAW_SECTOR_SIZE])
        if sha_bytes(sec) != w["after_sha256"]:
            raise SystemExit(f"FAIL Expected Write after SHA LBA {w['lba']}")
        if not verify_mode1_sector(sec)["valid"]:
            raise SystemExit(f"FAIL final EDC/ECC LBA {w['lba']}")

    asset_audit = []
    for r in asset_rows:
        try:
            name = str(r["asset"])
            lba = int(r["lba"])
            size = int(r["size"])
            target_sha = str(r["candidate_sha256"]).lower()
        except Exception as e:
            raise SystemExit("FAIL malformed Assets 58 row") from e
        final_sha = sha_bytes(extract_asset(out, lba, size))
        donor_asset_sha = sha_bytes(extract_asset(donor, lba, size))
        if donor_asset_sha != target_sha:
            raise SystemExit(f"FAIL donor whole-asset SHA {name}")
        if final_sha != target_sha:
            raise SystemExit(f"FAIL final whole-asset re-extraction {name}")
        asset_audit.append({
            "asset": name,
            "lba": lba,
            "size": size,
            "candidate_sha256": target_sha,
            "final_sha256": final_sha,
            "status": "PASS"
        })

    args.output_bin.parent.mkdir(parents=True, exist_ok=True)
    args.output_bin.write_bytes(out)
    output_sha = sha_file(args.output_bin)

    report = {
        "batch": 285,
        "status": "PASS_BATCH285_B284_PLUS_B118_STATIC58_SECTOR_UNION",
        "parent_batch": 284,
        "parent_sha256": parent_sha,
        "b118_donor_sha256": donor_sha,
        "b118_workbook_sha256": B118_WORKBOOK_SHA,
        "sector_audit": "1626/1626 PASS",
        "asset_reextraction": "58/58 PASS",
        "already_target_sectors": already_target,
        "expected_write_count": len(expected_write),
        "expected_write": expected_write,
        "changed_raw_sectors": len(actual_lbas),
        "changed_lbas": actual_lbas,
        "changed_sector_accounting": "PASS",
        "changed_sector_edc_ecc": f"{len(actual_lbas)}/{len(actual_lbas)} PASS",
        "asset_audit": asset_audit,
        "guessed_payload_bytes": 0,
        "output_sha256": output_sha,
        "scope": {
            "parent_reported_physical_assets": int(b284.get("physical_scope", {}).get("candidate_assets", 136)),
            "static_assets_verified": 58,
            "event_mes_logical_completion": b284.get("physical_scope", {}).get("event_mes_logical_completion", "109/109"),
            "note": "Physical asset union count is intentionally not arithmetically inflated because overlap is resolved by exact sector/asset gates."
        }
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(report["status"])
    print(f"output_sha256={output_sha}")
    print(f"changed_raw_sectors={len(actual_lbas)}")
    print("asset_reextraction=58/58 PASS")


if __name__ == "__main__":
    main()
