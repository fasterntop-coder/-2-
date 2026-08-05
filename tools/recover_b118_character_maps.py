#!/usr/bin/env python3
"""Recover exact SYSTEM/SYS14 character maps from the B118 Font Lifecycle sidecar.

Accepts the historical loose CSV export or workbook. A map is emitted only when
both bank rows pass structural, count, slot-range, uniqueness and reserved-slot
gates. No font or game bytes are included.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import tempfile
from pathlib import Path
from typing import Any, Iterable

NAME = "BATCH118_FONT_LIFECYCLE_MANIFEST.csv"
BANKS = ("SYSTEM", "SYS14")
EXPECTED_CUSTOM = {"SYSTEM": 364, "SYS14": 363}
SLOT_COUNT = 448
HEX = set("0123456789abcdef")


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        while block := stream.read(1024 * 1024):
            h.update(block)
    return h.hexdigest()


def norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    last: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                reader = csv.DictReader(stream)
                if not reader.fieldnames:
                    raise ValueError("missing CSV header")
                return [{norm(k).lower(): norm(v) for k, v in row.items()} for row in reader if any(norm(v) for v in row.values())]
        except UnicodeDecodeError as exc:
            last = exc
    raise ValueError(f"cannot decode CSV: {path}") from last


def looks_like_font_lifecycle(header: Iterable[str]) -> bool:
    names = {norm(value).lower() for value in header}
    return {"bank", "custom_characters", "reserved_slots", "character_map_json"}.issubset(names)


def workbook_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for workbook extraction") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        values = sheet.iter_rows(values_only=True)
        try:
            header = [norm(value).lower() for value in next(values)]
        except StopIteration:
            continue
        if not looks_like_font_lifecycle(header):
            continue
        rows: list[dict[str, str]] = []
        for values_row in values:
            row = {header[i]: norm(value) for i, value in enumerate(values_row) if i < len(header)}
            if any(row.values()):
                rows.append(row)
        return rows
    raise ValueError("Font Lifecycle sheet not found")


def find_input(root: Path) -> tuple[list[dict[str, str]], dict[str, str]]:
    loose = next(root.rglob(NAME), None)
    if loose:
        return read_csv(loose), {"mode": "loose_csv", "path": str(loose), "sha256": sha256(loose)}
    for workbook in root.rglob("*.xlsx"):
        try:
            rows = workbook_rows(workbook)
            return rows, {"mode": "workbook", "path": str(workbook), "sha256": sha256(workbook)}
        except Exception:
            continue
    raise FileNotFoundError("B118 Font Lifecycle CSV or workbook not found")


def parse_int(value: str, field: str, bank: str) -> int:
    try:
        return int(value, 0)
    except ValueError as exc:
        raise ValueError(f"{bank}: invalid {field}: {value!r}") from exc


def parse_reserved(value: str, bank: str) -> set[int]:
    value = value.strip()
    if not value:
        return set()
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        parsed = [part.strip() for part in value.replace(";", ",").split(",") if part.strip()]
    if isinstance(parsed, dict):
        parsed = list(parsed.values())
    if not isinstance(parsed, list):
        raise ValueError(f"{bank}: reserved_slots must be a JSON list or comma list")
    slots = {parse_int(str(item), "reserved slot", bank) for item in parsed}
    if any(not 0 <= slot < SLOT_COUNT for slot in slots):
        raise ValueError(f"{bank}: reserved slot outside 0..447")
    return slots


def valid_sha(value: str) -> bool:
    value = value.lower()
    return len(value) == 64 and all(ch in HEX for ch in value)


def validate(rows: list[dict[str, str]]) -> dict[str, dict[str, Any]]:
    selected = {norm(row.get("bank")).upper(): row for row in rows if norm(row.get("bank")).upper() in BANKS}
    if set(selected) != set(BANKS):
        raise ValueError(f"required banks missing or duplicated: found={sorted(selected)}")
    if sum(1 for row in rows if norm(row.get("bank")).upper() in BANKS) != 2:
        raise ValueError("Font Lifecycle must contain exactly one SYSTEM and one SYS14 row")
    result: dict[str, dict[str, Any]] = {}
    for bank in BANKS:
        row = selected[bank]
        custom = parse_int(norm(row.get("custom_characters")), "custom_characters", bank)
        if custom != EXPECTED_CUSTOM[bank]:
            raise ValueError(f"{bank}: custom_characters {custom} != {EXPECTED_CUSTOM[bank]}")
        try:
            mapping_raw = json.loads(norm(row.get("character_map_json")))
        except json.JSONDecodeError as exc:
            raise ValueError(f"{bank}: invalid character_map_json") from exc
        if not isinstance(mapping_raw, dict):
            raise ValueError(f"{bank}: character_map_json must be an object")
        mapping: dict[str, int] = {}
        for character, slot_value in mapping_raw.items():
            character = str(character)
            if len(character) != 1:
                raise ValueError(f"{bank}: key is not one Unicode character: {character!r}")
            slot = parse_int(str(slot_value), "character slot", bank)
            if not 0 <= slot < SLOT_COUNT:
                raise ValueError(f"{bank}: slot outside 0..447: {slot}")
            if character in mapping:
                raise ValueError(f"{bank}: duplicate character {character!r}")
            mapping[character] = slot
        if len(mapping) != custom:
            raise ValueError(f"{bank}: map entries {len(mapping)} != custom_characters {custom}")
        slots = list(mapping.values())
        if len(set(slots)) != len(slots):
            raise ValueError(f"{bank}: duplicate slot assignment")
        reserved = parse_reserved(norm(row.get("reserved_slots")), bank)
        collision = sorted(set(slots) & reserved)
        if collision:
            raise ValueError(f"{bank}: custom map collides with reserved slots: {collision}")
        font_sha = norm(row.get("font_source_sha256")).lower()
        if font_sha and not valid_sha(font_sha):
            raise ValueError(f"{bank}: invalid font_source_sha256")
        result[bank] = {
            "bank": bank,
            "slot_count": SLOT_COUNT,
            "custom_characters": custom,
            "reserved_slots": sorted(reserved),
            "font_source_sha256": font_sha,
            "character_to_slot": dict(sorted(mapping.items(), key=lambda item: item[1])),
        }
    return result


def recover(root: Path, output: Path) -> dict[str, Any]:
    rows, source = find_input(root)
    maps = validate(rows)
    output.mkdir(parents=True, exist_ok=True)
    outputs: dict[str, dict[str, str]] = {}
    for bank, document in maps.items():
        path = output / f"{bank}_CHARACTER_MAP.json"
        path.write_text(json.dumps(document, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        outputs[bank] = {"path": str(path), "sha256": sha256(path)}
    result = {
        "status": "PASS_B118_CHARACTER_MAPS_RECOVERED",
        "source": source,
        "banks": {bank: {"custom_characters": maps[bank]["custom_characters"], "reserved_slots": len(maps[bank]["reserved_slots"]), **outputs[bank]} for bank in BANKS},
    }
    (output / "B118_CHARACTER_MAP_RECOVERY_RESULT.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return result


def selftest() -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        path = root / NAME
        with path.open("w", encoding="utf-8", newline="") as stream:
            fields = ["bank", "font_source_sha256", "custom_characters", "reserved_slots", "character_map_json"]
            writer = csv.DictWriter(stream, fieldnames=fields)
            writer.writeheader()
            for bank in BANKS:
                count = EXPECTED_CUSTOM[bank]
                mapping = {chr(0xAC00 + index): index for index in range(count)}
                reserved = list(range(count, SLOT_COUNT))
                writer.writerow({
                    "bank": bank,
                    "font_source_sha256": "a" * 64,
                    "custom_characters": count,
                    "reserved_slots": json.dumps(reserved),
                    "character_map_json": json.dumps(mapping, ensure_ascii=False),
                })
        result = recover(root, root / "out")
        passed = result["status"] == "PASS_B118_CHARACTER_MAPS_RECOVERED"
        try:
            rows = read_csv(path)
            broken = json.loads(rows[0]["character_map_json"])
            second = next(iter(broken))
            broken[second] = 1
            rows[0]["character_map_json"] = json.dumps(broken, ensure_ascii=False)
            validate(rows)
            negative = False
        except ValueError:
            negative = True
        return {"status": "PASS" if passed and negative else "FAIL", "positive": passed, "negative_duplicate_slot": negative}


def main() -> int:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    run = sub.add_parser("recover")
    run.add_argument("search_root", type=Path)
    run.add_argument("--output-dir", type=Path, default=Path("output/B151_CHARACTER_MAPS"))
    sub.add_parser("selftest")
    args = parser.parse_args()
    result = selftest() if args.command == "selftest" else recover(args.search_root, args.output_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if str(result["status"]).startswith("PASS") else 2


if __name__ == "__main__":
    raise SystemExit(main())
