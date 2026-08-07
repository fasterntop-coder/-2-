#!/usr/bin/env python3
"""Extract and validate the exact 39-asset CD1 baseline from BATCH132_RECOVERY_STATUS.xlsx.

No guessed bytes are produced. This tool only accepts the historical workbook when all
asset geometry, SHA-256 strings, PASS states, aggregate counts, and verified candidate
BIN metadata are internally consistent.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook

HEX64 = re.compile(r"^[0-9a-f]{64}$")
EXPECTED_DISC_SIZE = 659_293_824
EXPECTED_ASSET_COUNT = 39
EXPECTED_BIN_SHA256 = "518c73a08e367a7f36c49a074ec7a91f61007456c99e4e30e73f8bb64575b250"
REQUIRED_B116 = {"SYS20", "SYS21", "SYS22", "SYS23", "SYS24", "SYS25", "SYS47", "STNSYS02", "STNSYS03"}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--output", type=Path, required=True)
    ns = ap.parse_args()

    wb = load_workbook(ns.workbook, data_only=True, read_only=True)
    if "39개 자산" not in wb.sheetnames or "배치132 요약" not in wb.sheetnames:
        raise SystemExit("BLOCKED: required B132 sheets are missing")

    ws = wb["39개 자산"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    expected_header = ["자산", "원천 배치", "LBA", "크기", "후보 SHA-256", "재추출"]
    if header[:6] != expected_header:
        raise SystemExit(f"BLOCKED: unexpected header {header[:6]!r}")

    assets = []
    names = set()
    extents = []
    for row in rows[1:]:
        if not row[0]:
            continue
        name, batch, lba, size, digest, status = row[:6]
        name = str(name)
        digest = str(digest).lower()
        if name in names:
            raise SystemExit(f"BLOCKED: duplicate asset {name}")
        if status != "PASS" or not HEX64.fullmatch(digest):
            raise SystemExit(f"BLOCKED: invalid record for {name}")
        lba, size, batch = int(lba), int(size), int(batch)
        sectors = (size + 2047) // 2048
        start, end = lba, lba + sectors
        if lba < 0 or size <= 0 or end * 2352 > EXPECTED_DISC_SIZE:
            raise SystemExit(f"BLOCKED: invalid geometry for {name}")
        names.add(name)
        extents.append((start, end, name))
        assets.append({"asset": name, "source_batch": batch, "lba": lba, "size": size, "sectors": sectors, "sha256": digest, "reextraction": "PASS"})

    if len(assets) != EXPECTED_ASSET_COUNT:
        raise SystemExit(f"BLOCKED: expected 39 assets, found {len(assets)}")
    if not REQUIRED_B116.issubset(names):
        raise SystemExit(f"BLOCKED: B116 exact banks missing: {sorted(REQUIRED_B116 - names)}")

    for (_, end_a, name_a), (start_b, _, name_b) in zip(sorted(extents), sorted(extents)[1:]):
        if end_a > start_b:
            raise SystemExit(f"BLOCKED: overlapping extents {name_a} / {name_b}")

    summary = wb["배치132 요약"]
    summary_rows = {str(r[0]): r[1] for r in summary.iter_rows(values_only=True) if r and r[0] is not None}
    if summary_rows.get("누적 정확 자산") != "39/58":
        raise SystemExit("BLOCKED: workbook aggregate is not 39/58")
    if summary_rows.get("재추출") != "39/39 PASS":
        raise SystemExit("BLOCKED: workbook re-extraction gate is not 39/39 PASS")
    if summary_rows.get("MODE1/2352 EDC/ECC") != "PASS":
        raise SystemExit("BLOCKED: workbook EDC/ECC gate failed")
    if summary_rows.get("검증 BIN SHA-256") != EXPECTED_BIN_SHA256:
        raise SystemExit("BLOCKED: unexpected historical candidate BIN SHA-256")

    result = {
        "format": "ST2-CD1-B132-EXACT39-v1",
        "status": "PASS_EXACT39_BASELINE_RECOVERED",
        "workbook": {"name": ns.workbook.name, "sha256": sha256_file(ns.workbook)},
        "source_disc": {"size": EXPECTED_DISC_SIZE, "format": "MODE1/2352"},
        "historical_verified_candidate": {
            "sha256": EXPECTED_BIN_SHA256,
            "changed_raw_sectors": 1134,
            "edc_ecc": "PASS",
            "reextraction": "39/39 PASS"
        },
        "coverage": {"exact_assets": 39, "total_assets": 58, "percent": 67.24137931034483},
        "assets": sorted(assets, key=lambda x: x["lba"]),
        "estimated_bytes": 0
    }
    ns.output.parent.mkdir(parents=True, exist_ok=True)
    ns.output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": result["status"], "assets": 39, "output": str(ns.output)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
