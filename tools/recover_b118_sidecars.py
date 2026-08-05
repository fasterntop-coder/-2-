#!/usr/bin/env python3
"""Recover and validate historical B118 Reverse Decode / Record Audit sidecars.

Loose CSV exports and the historical workbook are accepted. Legacy column names
are canonicalized before validation. Nothing is trusted by filename alone.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import tempfile
from collections import Counter
from pathlib import Path
from typing import Iterable

REVERSE_NAME = "BATCH118_REVERSE_DECODE.csv"
AUDIT_NAME = "BATCH118_RECORD_AUDIT_458.csv"
REVERSE_ROWS = 445
AUDIT_ROWS = 458
REVERSE_SPLIT = {"SYSTEM": 222, "SYS14": 223}
AUDIT_SPLIT = {"SYSTEM": 229, "SYS14": 229}
HEX = set("0123456789abcdef")

REVERSE_ALIASES = {
    "bank": ("bank",),
    "record": ("record", "record_index", "index"),
    "expected": ("expected", "expected_korean", "expected_text"),
    "decoded": ("decoded", "decoded_korean", "decoded_text"),
    "status": ("status", "result"),
    "type": ("type",),
}
AUDIT_ALIASES = {
    "bank": ("bank",),
    "record": ("record", "record_index", "index"),
    "source_record_sha256": ("source_record_sha256", "source_sha256"),
    "candidate_record_sha256": ("candidate_record_sha256", "candidate_sha256"),
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as stream:
        while block := stream.read(1024 * 1024):
            h.update(block)
    return h.hexdigest()


def norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    last: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                reader = csv.DictReader(stream)
                if not reader.fieldnames:
                    raise ValueError(f"missing CSV header: {path}")
                return [{norm(k).lower(): norm(v) for k, v in row.items()} for row in reader]
        except UnicodeDecodeError as exc:
            last = exc
    raise ValueError(f"cannot decode CSV: {path}") from last


def canonicalize(row: dict[str, str], aliases: dict[str, tuple[str, ...]]) -> dict[str, str]:
    lowered = {norm(k).lower(): norm(v) for k, v in row.items()}
    result = dict(lowered)
    for canonical, names in aliases.items():
        for name in names:
            if lowered.get(name, ""):
                result[canonical] = lowered[name]
                break
        else:
            result.setdefault(canonical, "")
    return result


def classify_header(header: Iterable[str]) -> str | None:
    names = {norm(value).lower() for value in header}
    reverse_signals = {"expected", "decoded"} | {"decoded_korean", "expected_korean"}
    if "bank" in names and ("record" in names or "index" in names) and len(names & reverse_signals) >= 2:
        return "reverse"
    if "bank" in names and ("record" in names or "index" in names) and {
        "source_record_sha256", "candidate_record_sha256"
    }.issubset(names):
        return "audit"
    return None


def workbook_rows(path: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for workbook extraction") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    found: dict[str, list[dict[str, str]]] = {}
    for sheet in workbook.worksheets:
        values = sheet.iter_rows(values_only=True)
        try:
            header = [norm(value).lower() for value in next(values)]
        except StopIteration:
            continue
        kind = classify_header(header)
        if not kind:
            continue
        rows = []
        for values_row in values:
            row = {header[index]: norm(value) for index, value in enumerate(values_row) if index < len(header)}
            if any(row.values()):
                rows.append(row)
        found[kind] = rows
    if "reverse" not in found or "audit" not in found:
        raise ValueError("required Reverse Decode / Record Audit sheets not found")
    return found["reverse"], found["audit"]


def find_inputs(root: Path) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, str]]:
    reverse = next(root.rglob(REVERSE_NAME), None)
    audit = next(root.rglob(AUDIT_NAME), None)
    if reverse and audit:
        return read_csv(reverse), read_csv(audit), {
            "mode": "loose_csv", "reverse": str(reverse), "audit": str(audit),
            "reverse_sha256": sha256(reverse), "audit_sha256": sha256(audit),
        }
    for workbook in root.rglob("*.xlsx"):
        try:
            reverse_rows, audit_rows = workbook_rows(workbook)
            return reverse_rows, audit_rows, {
                "mode": "workbook", "workbook": str(workbook), "workbook_sha256": sha256(workbook)
            }
        except Exception:
            continue
    raise FileNotFoundError("B118 sidecar CSVs or workbook not found")


def parse_record(value: str, kind: str) -> int:
    try:
        return int(value)
    except ValueError as exc:
        raise ValueError(f"{kind}: invalid record index {value!r}") from exc


def validate_reverse(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized = [canonicalize(row, REVERSE_ALIASES) for row in rows]
    if len(normalized) != REVERSE_ROWS:
        raise ValueError(f"reverse: rows {len(normalized)} != {REVERSE_ROWS}")
    counts = Counter(row["bank"] for row in normalized)
    if dict(counts) != REVERSE_SPLIT:
        raise ValueError(f"reverse: bank split {dict(counts)} != {REVERSE_SPLIT}")
    seen: set[tuple[str, int]] = set()
    for row in normalized:
        record = parse_record(row["record"], "reverse")
        key = (row["bank"], record)
        if key in seen or not 0 <= record <= 228:
            raise ValueError(f"reverse: invalid or duplicate key {key}")
        seen.add(key)
        expected, decoded = row["expected"], row["decoded"]
        if not expected or not decoded:
            raise ValueError(f"reverse: blank expected/decoded at {key}")
        if expected != decoded:
            raise ValueError(f"reverse: expected/decoded mismatch at {key}")
        if row.get("status") and row["status"].upper() != "PASS":
            raise ValueError(f"reverse: non-PASS status at {key}: {row['status']}")
        row["record"] = str(record)
        row["decoded_korean"] = decoded
    return normalized


def valid_sha(value: str) -> bool:
    value = value.lower()
    return len(value) == 64 and all(character in HEX for character in value)


def validate_audit(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized = [canonicalize(row, AUDIT_ALIASES) for row in rows]
    if len(normalized) != AUDIT_ROWS:
        raise ValueError(f"audit: rows {len(normalized)} != {AUDIT_ROWS}")
    counts = Counter(row["bank"] for row in normalized)
    if dict(counts) != AUDIT_SPLIT:
        raise ValueError(f"audit: bank split {dict(counts)} != {AUDIT_SPLIT}")
    seen: set[tuple[str, int]] = set()
    per_bank: dict[str, set[int]] = {"SYSTEM": set(), "SYS14": set()}
    for row in normalized:
        record = parse_record(row["record"], "audit")
        key = (row["bank"], record)
        if key in seen or not 0 <= record <= 228:
            raise ValueError(f"audit: invalid or duplicate key {key}")
        seen.add(key)
        per_bank.setdefault(row["bank"], set()).add(record)
        for field in ("source_record_sha256", "candidate_record_sha256"):
            if not valid_sha(row[field]):
                raise ValueError(f"audit: invalid {field} at {key}")
        row["record"] = str(record)
    expected = set(range(229))
    for bank in ("SYSTEM", "SYS14"):
        if per_bank.get(bank) != expected:
            missing = sorted(expected - per_bank.get(bank, set()))
            raise ValueError(f"audit: incomplete {bank} coverage; missing={missing}")
    return normalized


def write_csv(path: Path, rows: list[dict[str, str]], preferred: list[str]) -> str:
    fields = preferred + sorted({key for row in rows for key in row} - set(preferred))
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return sha256(path)


def recover(root: Path, output: Path) -> dict[str, object]:
    reverse_raw, audit_raw, source = find_inputs(root)
    reverse = validate_reverse(reverse_raw)
    audit = validate_audit(audit_raw)
    audit_keys = {(row["bank"], int(row["record"])) for row in audit}
    reverse_keys = {(row["bank"], int(row["record"])) for row in reverse}
    if not reverse_keys.issubset(audit_keys):
        raise ValueError("reverse rows are not a subset of audited records")
    output.mkdir(parents=True, exist_ok=True)
    reverse_path, audit_path = output / REVERSE_NAME, output / AUDIT_NAME
    result = {
        "status": "PASS_B118_SIDECARS_RECOVERED",
        "source": source,
        "reverse_rows": len(reverse),
        "audit_rows": len(audit),
        "reverse_bank_split": dict(Counter(row["bank"] for row in reverse)),
        "audit_bank_split": dict(Counter(row["bank"] for row in audit)),
        "reverse_sha256": write_csv(reverse_path, reverse, ["bank", "record", "type", "expected", "decoded", "status", "decoded_korean"]),
        "audit_sha256": write_csv(audit_path, audit, ["bank", "record", "type", "source", "capacity", "compiled", "margin", "metadata_exact", "linebreak_exact", "source_record_sha256", "candidate_record_sha256", "expected_write"]),
        "reverse_path": str(reverse_path),
        "audit_path": str(audit_path),
    }
    (output / "B118_SIDECAR_RECOVERY_RESULT.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return result


def selftest() -> dict[str, object]:
    with tempfile.TemporaryDirectory() as directory:
        root, output = Path(directory), Path(directory) / "output"
        with (root / REVERSE_NAME).open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=["index", "bank", "record", "type", "expected", "decoded", "status"])
            writer.writeheader()
            for index in range(REVERSE_ROWS):
                bank = "SYSTEM" if index < 222 else "SYS14"
                record = index if index < 222 else index - 222
                text = f"문장{index}"
                writer.writerow({"index": index, "bank": bank, "record": record, "type": "UNIQUE", "expected": text, "decoded": text, "status": "PASS"})
        with (root / AUDIT_NAME).open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=["index", "bank", "record", "source_record_sha256", "candidate_record_sha256"])
            writer.writeheader()
            for index in range(AUDIT_ROWS):
                bank = "SYSTEM" if index < 229 else "SYS14"
                record = index if index < 229 else index - 229
                writer.writerow({
                    "index": index, "bank": bank, "record": record,
                    "source_record_sha256": hashlib.sha256(f"s{index}".encode()).hexdigest(),
                    "candidate_record_sha256": hashlib.sha256(f"c{index}".encode()).hexdigest(),
                })
        result = recover(root, output)
        passed = result["status"] == "PASS_B118_SIDECARS_RECOVERED"
        return {"status": "PASS" if passed else "FAIL", "legacy_schema": passed}


def main() -> int:
    parser = argparse.ArgumentParser()
    commands = parser.add_subparsers(dest="command", required=True)
    run = commands.add_parser("recover")
    run.add_argument("root", type=Path)
    run.add_argument("--output-dir", type=Path, default=Path("output/B118_SIDECARS"))
    commands.add_parser("selftest")
    args = parser.parse_args()
    result = selftest() if args.command == "selftest" else recover(args.root, args.output_dir)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if str(result["status"]).startswith("PASS") else 2


if __name__ == "__main__":
    raise SystemExit(main())
